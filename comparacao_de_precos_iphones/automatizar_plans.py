import openpyxl


def create_sheet():
    # Cria um novo Workbook
    book = openpyxl.Workbook()
    print(book.sheetnames) # Acessa os nomes da planilhas
    return book

# Cria o workbook
book = create_sheet()

# Cria uma nova planilha chamada 'Iphones'
book.create_sheet('Iphones')

# Acessa a planilha 'Iphones'
iphones_page = book['Iphones']

# Adiciona os cabeçalhos à planilha
iphones_page.append(['Índice', 'Tipo', 'Preço Loja 1', 'Preço Loja 2', 'Preço Loja 3'])
iphones_page.append(['1', 'Iphone 11', 'R$2.500,00', 'R$2.654,00', 'R$3.200,00'])
iphones_page.append(['2', 'Iphone 12', 'R$2.800,00', 'R$3.400,00', 'R$3.560,00'])
iphones_page.append(['3', 'Iphone 13', 'R$4.000,00', 'R$4.200,00', 'R$4.350,00'])

# Salva o workbook em um arquivo
book.save('comparacao_precos_iphones.xlsx')